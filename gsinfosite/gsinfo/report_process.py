# encoding=UTF-8
import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from .models import *
from gsinfosite import settings
import time
from math import ceil
import zipfile
from openpyxl.utils.units import (
    DEFAULT_ROW_HEIGHT,
    DEFAULT_COLUMN_WIDTH
)
from openpyxl.utils.units import points_to_pixels
import re
import math
from decimal import *

def getRows(value_list):
    # 判断value中的每一个值是中文还是英文
    zhPattern = re.compile(u'[\u4e00-\u9fa5]')
    row_list = []
    for value in value_list:
        vs,n = value
        all_len = 0
        for v in unicode(str(vs), "utf-8"):
            match = zhPattern.search(v)
            if match:
                all_len += 2
            else:
                all_len += 1
        row_num = int(math.ceil(float(all_len)/float(n)))
        row_list.append(row_num)
    max_row = max(row_list)
    return max_row

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def createBoxInfo(**kwargs):
    boxNumber = kwargs['boxNumber']

    boxDir = os.path.join(settings.BOX_DATA_PATH, boxNumber)
    if not os.path.exists(boxDir):
        os.mkdir(boxDir)  # 转移至创建作业时, 生成相应的目录
    box = gsBox.objects.get(boxNumber=boxNumber)
    thing_set = gsThing.objects.filter(box=box)
    prop = box.boxType
    type = prop.type
    parentType = prop.parentType
    grandpaType = prop.grandpaType
    if grandpaType:
        className = parentType
        subClassName = type
    else:
        className = type
        subClassName = '-'
    template_path = os.path.join(settings.TEMPLATE_DATA_PATH, u'装箱清单.xlsx')
    table = load_workbook(template_path)
    # default_height = points_to_pixels(DEFAULT_ROW_HEIGHT)
    sheet = table.worksheets[0]
    # default_height=20,而只有高度大于默认值的单元格sheet.row_dimensions[n].height才不返回None
    unit = 14.25
    if not sheet.row_dimensions[1].height:
        sheet.row_dimensions[1].height = unit
        h1 = sheet.row_dimensions[1].height
    else:
        h1 = sheet.row_dimensions[1].height

    if not sheet.row_dimensions[2].height:
        sheet.row_dimensions[2].height = unit
        h2 = sheet.row_dimensions[2].height
    else:
        h2 = sheet.row_dimensions[2].height
    if not sheet.row_dimensions[3].height:
        sheet.row_dimensions[3].height = unit
        h3 = sheet.row_dimensions[3].height
    else:
        h3 = sheet.row_dimensions[3].height
    h = h1 + h2 + h3
    print_area = 34 * unit
    name_count=1
    row_count=1
    filePath_list = []
    row_height_list = []
    rowStartIdx = 4

    # 实线边框样式
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    outline=Side(border_style='thin', color='FF000000'),
                    vertical=Side(border_style='thin', color='FF000000'),
                    horizontal=Side(border_style='thin', color='FF000000')
                    )
    # font = Font(name=u'仿宋', size=10)
    # alignment = Alignment(horizontal='center', vertical='center')
    # 写合计用
    grossWeight_heji = []
    detectedQuantity_heji = []
    pureWeight_heji = []
    amount_heji =[]
    # 写小计用
    grossWeight_xiaoji = []
    detectedQuantity_xiaoji = []
    pureWeight_xiaoji = []
    amount_xiaoji = []

    page = 0 # 只有一页的话就只写合计不写小计
    for i,th in enumerate(thing_set):
        # 写表头
        value_list = []
        sheet['A{0}'.format(2)] = u'箱号:' + boxNumber
        date = datetime.datetime.now()
        year = date.year
        month = date.month
        day = date.day
        sheet['G{0}'.format(2)] = u'{0}年{1}月{2}日'.format(year, month, day)
        # 写表
        # 序号
        col1 = i + 1
        sheet.cell(row=rowStartIdx, column=1).value = col1
        value_list.append((col1,4))
        # 盒号
        col2 = th.case.caseNumber if th.case else '-'
        sheet.cell(row=rowStartIdx, column=2).value = col2
        value_list.append((col2,24))
        # 实物编号
        col3 = th.serialNumber2 if th.serialNumber2 else '-'
        sheet.cell(row=rowStartIdx, column=3).value = col3
        value_list.append((col3, 24))
        # 品种
        col4= className
        sheet.cell(row=rowStartIdx, column=4).value = col4
        value_list.append((col4,12))
        # 品名
        col5 =subClassName
        sheet.cell(row=rowStartIdx, column=5).value =col5
        value_list.append((col5,8))
        # 等级
        col6 =th.level
        sheet.cell(row=rowStartIdx, column=6).value = col6
        value_list.append((col6,6))
        # 名称
        col7 =th.detailedName
        sheet.cell(row=rowStartIdx, column=7).value = col7
        value_list.append((col7,28))
        # 毛重
        col8 = '%.2f' % th.grossWeight if th.grossWeight else '%.2f' % 0
        value_len = len(col8)
        if value_len > 9:
            sheet.cell(row=rowStartIdx, column=8).value = col8
        else:
            sheet.cell(row=rowStartIdx, column=8).value =Decimal(col8).quantize(Decimal('0.00'))
        grossWeight_xiaoji.append(float(col8))
        # 检测成色
        col9 = '%.2f' % th.detectedQuantity if th.detectedQuantity else '%.2f' % 0
        value_len = len(col9)
        if value_len >9:
            sheet.cell(row=rowStartIdx, column=9).value = col9
        else:
            sheet.cell(row=rowStartIdx, column=9).value = Decimal(col9).quantize(Decimal('0.00'))
        detectedQuantity_xiaoji.append(float(col9))
        # 纯重
        col10 = '%.2f' % th.pureWeight if th.pureWeight else '%.2f' % 0
        value_len = len(col10)
        if value_len > 9:
            sheet.cell(row=rowStartIdx, column=10).value =col10
        else:
            sheet.cell(row=rowStartIdx, column=10).value = Decimal(col10).quantize(Decimal('0.00'))
        pureWeight_xiaoji.append(float(col10))
        # 件（枚）数
        col11 = th.amount
        sheet.cell(row=rowStartIdx, column=11).value =col11
        amount_xiaoji.append(float(col11))
        # ----------------------------
        max_row = getRows(value_list)
        sheet.row_dimensions[rowStartIdx].height = unit * max_row
        h_some = unit * max_row
        row_height_list.append(h_some)
        sum_row_height = sum(row_height_list)
        real_h = h + sum_row_height
        other_area = print_area - real_h
        if other_area <= 3 * unit:
            page += 1
            # 写小计
            sheet.cell(row=rowStartIdx + 1, column=1).value = u'小计'
            sheet.cell(row=rowStartIdx + 1, column=2).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=3).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=4).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=5).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=6).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=7).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=8).value = Decimal(str(sum(grossWeight_xiaoji))).quantize(Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=9).value = Decimal(str(sum(detectedQuantity_xiaoji))).quantize(Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=10).value = Decimal(str(sum(pureWeight_xiaoji))).quantize(Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=11).value = Decimal(str(sum(amount_xiaoji))).quantize(Decimal('0.00'))
            # 写签名
            sheet.merge_cells('A{0}:C{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet.merge_cells('D{0}:G{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet.merge_cells('H{0}:K{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet['A{0}'.format(rowStartIdx + 2)] = '现场负责人：'
            sheet['A{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet['D{0}'.format(rowStartIdx + 2)] = '复核：'
            sheet['D{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet['H{0}'.format(rowStartIdx + 2)] = '制单：'
            sheet['H{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet.row_dimensions[rowStartIdx + 2].height = 2 * unit
            # 写样式
            sub_area1 = sheet['A{0}:K{1}'.format(3, rowStartIdx+1)]
            for cs in sub_area1:
                for cell in cs:
                    cell.border = border
            # 保存
            fileName = u'{0}-{1}.xlsx'.format(boxNumber,str(name_count))
            filePath = os.path.join(boxDir, fileName)
            filePath_list.append(filePath)
            table.save(filePath)
            # -------------
            grossWeight_heji = grossWeight_heji + grossWeight_xiaoji
            detectedQuantity_heji = detectedQuantity_heji +detectedQuantity_xiaoji
            pureWeight_heji = pureWeight_heji + pureWeight_xiaoji
            amount_heji = amount_heji + amount_xiaoji
            # 重新初始化
            name_count += 1
            rowStartIdx = 4
            row_height_list = []
            template_path = os.path.join(settings.TEMPLATE_DATA_PATH, u'装箱清单.xlsx')
            table = load_workbook(template_path)
            sheet = table.worksheets[0]

            grossWeight_xiaoji = []
            detectedQuantity_xiaoji = []
            pureWeight_xiaoji = []
            amount_xiaoji = []

        else:
            rowStartIdx +=1
            row_count +=1
    if page > 0:
        # 写小计
        sheet.cell(row=rowStartIdx, column=1).value = u'小计'
        sheet.cell(row=rowStartIdx, column=2).value = '-'
        sheet.cell(row=rowStartIdx, column=3).value = '-'
        sheet.cell(row=rowStartIdx, column=4).value = '-'
        sheet.cell(row=rowStartIdx, column=5).value = '-'
        sheet.cell(row=rowStartIdx, column=6).value = '-'
        sheet.cell(row=rowStartIdx, column=7).value = '-'
        sheet.cell(row=rowStartIdx, column=8).value = Decimal(str(sum(grossWeight_xiaoji))).quantize(Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=9).value = Decimal(str(sum(detectedQuantity_xiaoji))).quantize(Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=10).value = Decimal(str(sum(pureWeight_xiaoji))).quantize(Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=11).value = Decimal(str(sum(amount_xiaoji))).quantize(Decimal('0.00'))

        # 以下方式(写公式)的话会导致excel save后，手动打开保存的excel在未作任何操作的情况下一开始就处于修改状态
        # sheet.cell(row=rowStartIdx, column=8).value = u'=SUM(H4:H{0})'.format(rowStartIdx-1)
        # sheet.cell(row=rowStartIdx, column=9).value = u'=SUM(I4:I{0})'.format(rowStartIdx-1)
        # sheet.cell(row=rowStartIdx, column=10).value = u'=SUM(J4:J{0})'.format(rowStartIdx-1)
        # sheet.cell(row=rowStartIdx, column=11).value = u'=SUM(K4:K{0})'.format(rowStartIdx-1)

    # 写合计
    sheet['A{0}'.format(rowStartIdx + 1)] = u'合计'
    sheet['B{0}'.format(rowStartIdx + 1)] = '-'
    sheet['C{0}'.format(rowStartIdx + 1)] = '-'
    sheet['D{0}'.format(rowStartIdx + 1)] = '-'
    sheet['E{0}'.format(rowStartIdx + 1)] = '-'
    sheet['F{0}'.format(rowStartIdx + 1)] = '-'
    sheet['G{0}'.format(rowStartIdx + 1)] = '-'

    sheet['H{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(grossWeight_heji))).quantize(Decimal('0.00'))
    sheet['I{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(detectedQuantity_heji))).quantize(Decimal('0.00'))
    sheet['J{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(pureWeight_heji))).quantize(Decimal('0.00'))
    sheet['K{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(amount_heji))).quantize(Decimal('0.00'))
    # 写签名
    sheet.merge_cells('A{0}:C{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet.merge_cells('D{0}:G{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet.merge_cells('H{0}:K{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet['A{0}'.format(rowStartIdx + 2)] = '现场负责人：'
    sheet['A{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet['D{0}'.format(rowStartIdx + 2)] = '复核：'
    sheet['D{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet['H{0}'.format(rowStartIdx + 2)] = '制单：'
    sheet['H{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet.row_dimensions[rowStartIdx + 2].height = 2 * unit
    # 写样式
    sub_area1 = sheet['A{0}:K{1}'.format(3, rowStartIdx + 1)]
    for cs in sub_area1:
        for cell in cs:
            cell.border = border
    # 保存
    fileName = u'{0}-{1}.xlsx'.format(boxNumber,str(name_count))
    filePath = os.path.join(boxDir, fileName)
    filePath_list.append(filePath)
    table.save(filePath)
    all_file_path = '|'.join(filePath_list)
    return all_file_path
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
'''
>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> sheet['A1'] = 'Tall row'
>>> sheet['B2'] = 'Wide column'
>>> sheet.row_dimensions[1].height = 70
>>> sheet.column_dimensions['B'].width = 20
>>> wb.save('dimensions.xlsx')
'''
# --------------------------------------装盒票--------------------------------------
def createCaseTicket(**kwargs):
    boxNumber = kwargs['boxNumber']
    caseNumber = kwargs['caseNumber']
    serialNumber2_list = kwargs['serialNumber2_list']

    boxDir = os.path.join(settings.BOX_DATA_PATH, boxNumber)
    if not os.path.exists(boxDir):
        os.mkdir(boxDir)

    box = gsBox.objects.get(boxNumber=boxNumber)
    case = gsCase.objects.get(caseNumber=caseNumber)
    thing_set = gsThing.objects.filter(serialNumber2__in=serialNumber2_list,case=case)

    prop = box.boxType
    type = prop.type
    parentType = prop.parentType
    grandpaType = prop.grandpaType
    if grandpaType:
        className = parentType
        subClassName = type
    else:
        className = type
        subClassName = '-'
    template_path = os.path.join(settings.TEMPLATE_DATA_PATH, u'装盒票.xlsx')
    table = load_workbook(template_path)
    sheet = table.worksheets[0]
    unit = 14.25
    if not sheet.row_dimensions[1].height:
        sheet.row_dimensions[1].height = unit
        h1 = sheet.row_dimensions[1].height
    else:
        h1 = sheet.row_dimensions[1].height

    if not sheet.row_dimensions[2].height:
        sheet.row_dimensions[2].height = unit
        h2 = sheet.row_dimensions[2].height
    else:
        h2 = sheet.row_dimensions[2].height
    if not sheet.row_dimensions[3].height:
        sheet.row_dimensions[3].height = unit
        h3 = sheet.row_dimensions[3].height
    else:
        h3 = sheet.row_dimensions[3].height
    h = h1 + h2 + h3
    print_area = 34 * unit  # 打印区域
    name_count = 1
    row_count = 1
    filePath_list = []
    row_height_list = []
    rowStartIdx = 4

    # 实线边框样式
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    outline=Side(border_style='thin', color='FF000000'),
                    vertical=Side(border_style='thin', color='FF000000'),
                    horizontal=Side(border_style='thin', color='FF000000')
                    )
    # font = Font(name=u'仿宋', size=10)
    # alignment = Alignment(horizontal='center', vertical='center')
    # 写合计用
    grossWeight_heji = []
    detectedQuantity_heji = []
    pureWeight_heji = []
    amount_heji = []
    # 写小计用
    grossWeight_xiaoji = []
    detectedQuantity_xiaoji = []
    pureWeight_xiaoji = []
    amount_xiaoji = []

    page = 0  # 只有一页的话就只写合计不写小计
    for i, th in enumerate(thing_set):
        # 写表头
        value_list = []
        sheet['A{0}'.format(2)] = u'盒号:' + caseNumber
        date = datetime.datetime.now()
        year = date.year
        month = date.month
        day = date.day
        sheet['G{0}'.format(2)] = u'{0}年{1}月{2}日'.format(year, month, day)
        # 写表
        # 序号
        col1 = i + 1
        sheet.cell(row=rowStartIdx, column=1).value = col1
        value_list.append((col1, 4))
        # 实物编号
        col2 = th.serialNumber2
        sheet.cell(row=rowStartIdx, column=2).value = col2
        value_list.append((col2, 25))
        # 品种
        col3 = className
        sheet.cell(row=rowStartIdx, column=3).value = col3
        value_list.append((col3, 14))
        # 品名
        col4 = subClassName
        sheet.cell(row=rowStartIdx, column=4).value = col4
        value_list.append((col4, 10))
        # 等级
        col5 = th.level
        sheet.cell(row=rowStartIdx, column=5).value = col5
        value_list.append((col5, 8))
        # 名称
        col6 = th.detailedName
        sheet.cell(row=rowStartIdx, column=6).value = col6
        value_list.append((col6, 32))
        # 毛重
        col7 = '%.2f' % th.grossWeight if th.grossWeight else '%.2f' % 0
        value_len = len(col7)
        if value_len > 12:
            sheet.cell(row=rowStartIdx, column=7).value = col7
        else:
            sheet.cell(row=rowStartIdx, column=7).value = Decimal(col7).quantize(Decimal('0.00'))
        grossWeight_xiaoji.append(float(col7))
        # 检测成色
        col8 = '%.2f' % th.detectedQuantity if th.detectedQuantity else '%.2f' % 0
        value_len = len(col8)
        if value_len > 12:
            sheet.cell(row=rowStartIdx, column=8).value = col8
        else:
            sheet.cell(row=rowStartIdx, column=8).value = Decimal(col8).quantize(Decimal('0.00'))
        detectedQuantity_xiaoji.append(float(col8))
        # 纯重
        col9 = '%.2f' % th.pureWeight if th.pureWeight else '%.2f' % 0
        value_len = len(col9)
        if value_len > 12:
            sheet.cell(row=rowStartIdx, column=9).value = col9
        else:
            sheet.cell(row=rowStartIdx, column=9).value = Decimal(col9).quantize(Decimal('0.00'))
        pureWeight_xiaoji.append(float(col9))
        # 件（枚）数
        col10 = th.amount
        sheet.cell(row=rowStartIdx, column=10).value = col10
        amount_xiaoji.append(float(col10))
        # ----------------------------
        max_row = getRows(value_list)
        sheet.row_dimensions[rowStartIdx].height = unit * max_row
        h_some = unit * max_row
        row_height_list.append(h_some)
        sum_row_height = sum(row_height_list)
        real_h = h + sum_row_height
        other_area = print_area - real_h
        if other_area <= 3 * unit:
            page += 1
            # 写小计
            sheet.cell(row=rowStartIdx + 1, column=1).value = u'小计'
            sheet.cell(row=rowStartIdx + 1, column=2).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=3).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=4).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=5).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=6).value = '-'
            sheet.cell(row=rowStartIdx + 1, column=7).value = Decimal(str(sum(grossWeight_xiaoji))).quantize(
                Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=8).value = Decimal(str(sum(detectedQuantity_xiaoji))).quantize(
                Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=9).value = Decimal(str(sum(pureWeight_xiaoji))).quantize(
                Decimal('0.00'))
            sheet.cell(row=rowStartIdx + 1, column=10).value = Decimal(str(sum(amount_xiaoji))).quantize(
                Decimal('0.00'))
            # 写签名
            sheet.merge_cells('A{0}:C{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet.merge_cells('D{0}:F{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet.merge_cells('G{0}:J{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
            sheet['A{0}'.format(rowStartIdx + 2)] = '现场负责人：'
            sheet['A{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet['D{0}'.format(rowStartIdx + 2)] = '复核：'
            sheet['D{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet['G{0}'.format(rowStartIdx + 2)] = '制单：'
            sheet['G{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
            sheet.row_dimensions[rowStartIdx + 2].height = 2 * unit
            # 写样式
            sub_area1 = sheet['A{0}:J{1}'.format(3, rowStartIdx + 1)]
            for cs in sub_area1:
                for cell in cs:
                    cell.border = border
            # 保存
            fileName = u'{0}-{1}.xlsx'.format(caseNumber,str(name_count))
            filePath = os.path.join(boxDir, fileName)
            filePath_list.append(filePath)
            table.save(filePath)
            # -------------
            grossWeight_heji = grossWeight_heji + grossWeight_xiaoji
            detectedQuantity_heji = detectedQuantity_heji + detectedQuantity_xiaoji
            pureWeight_heji = pureWeight_heji + pureWeight_xiaoji
            amount_heji = amount_heji + amount_xiaoji
            # 重新初始化
            name_count += 1
            rowStartIdx = 4
            row_height_list = []
            template_path = os.path.join(settings.TEMPLATE_DATA_PATH, u'装盒票.xlsx')
            table = load_workbook(template_path)
            sheet = table.worksheets[0]

            grossWeight_xiaoji = []
            detectedQuantity_xiaoji = []
            pureWeight_xiaoji = []
            amount_xiaoji = []

        else:
            rowStartIdx += 1
            row_count += 1
    if page > 0:
        # 写小计
        sheet.cell(row=rowStartIdx, column=1).value = u'小计'
        sheet.cell(row=rowStartIdx, column=2).value = '-'
        sheet.cell(row=rowStartIdx, column=3).value = '-'
        sheet.cell(row=rowStartIdx, column=4).value = '-'
        sheet.cell(row=rowStartIdx, column=5).value = '-'
        sheet.cell(row=rowStartIdx, column=6).value = '-'
        sheet.cell(row=rowStartIdx, column=7).value = Decimal(str(sum(grossWeight_xiaoji))).quantize(Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=8).value = Decimal(str(sum(detectedQuantity_xiaoji))).quantize(
            Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=9).value = Decimal(str(sum(pureWeight_xiaoji))).quantize(Decimal('0.00'))
        sheet.cell(row=rowStartIdx, column=10).value = Decimal(str(sum(amount_xiaoji))).quantize(Decimal('0.00'))
    # 写合计
    sheet['A{0}'.format(rowStartIdx + 1)] = u'合计'
    sheet['B{0}'.format(rowStartIdx + 1)] = '-'
    sheet['C{0}'.format(rowStartIdx + 1)] = '-'
    sheet['D{0}'.format(rowStartIdx + 1)] = '-'
    sheet['E{0}'.format(rowStartIdx + 1)] = '-'
    sheet['F{0}'.format(rowStartIdx + 1)] = '-'
    sheet['G{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(grossWeight_heji))).quantize(Decimal('0.00'))
    sheet['H{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(detectedQuantity_heji))).quantize(Decimal('0.00'))
    sheet['I{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(pureWeight_heji))).quantize(Decimal('0.00'))
    sheet['J{0}'.format(rowStartIdx + 1)] = Decimal(str(sum(amount_heji))).quantize(Decimal('0.00'))
    # 写签名
    sheet.merge_cells('A{0}:C{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet.merge_cells('D{0}:F{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet.merge_cells('G{0}:J{0}'.format(rowStartIdx + 2, rowStartIdx + 2))
    sheet['A{0}'.format(rowStartIdx + 2)] = '现场负责人：'
    sheet['A{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet['D{0}'.format(rowStartIdx + 2)] = '复核：'
    sheet['D{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet['G{0}'.format(rowStartIdx + 2)] = '制单：'
    sheet['G{0}'.format(rowStartIdx + 2)].alignment = Alignment(horizontal="left")
    sheet.row_dimensions[rowStartIdx + 2].height = 2 * unit
    # 写样式
    sub_area1 = sheet['A{0}:J{1}'.format(3, rowStartIdx + 1)]
    for cs in sub_area1:
        for cell in cs:
            cell.border = border
    # 保存
    fileName = u'{0}-{1}.xlsx'.format(caseNumber,str(name_count))
    filePath = os.path.join(boxDir, fileName)
    filePath_list.append(filePath)
    table.save(filePath)
    all_file_path = '|'.join(filePath_list)
    return all_file_path