# encoding=UTF-8
import os
import qrcode  # 二维码生成库
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from models import *
from gsinfosite import settings
import utils
import time

tagRootDir = settings.DATA_DIRS['tag_dir']
tagDir = tagRootDir
TempDir = tagDir


def createQRCode(c):
    '''
    :param c: 二维码图片标记
    :return: 
    '''
    img = qrcode.make(c)
    fileName = str(c) + '.png'
    filePath = os.path.join(tempDir, fileName)
    img.resize((128, 128)).save(filePath)
    return filePath


def createBarCode(c):
    pass


def writeDataIntoXLS(ws, pic_path, c, row, col, seq):
    '''
    :param ws: 
    :param pic_path: 
    :param c: tag图片名称
    :param row: 
    :param col: 
    :param seq: 
    :return: 
    '''
    idx = seq % 9
    # QRCode picture and Text 
    font = Font(name=u'Times New Roman', size=13, bold=True)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if (idx < 3):
        if (0 == idx % 3):
            pic_pos_row = row
            pic_pos_col = col
            text_pos_left_row = row + 1
            text_pos_left_col = chr(ord(col) + 1)
            text_pos_right_row = row + 2 + 3
            text_pos_right_col = chr(ord(col) + 2)
        elif (1 == idx % 3):
            pic_pos_row = row
            pic_pos_col = chr(ord(col) + 4)
            text_pos_left_row = row + 1
            text_pos_left_col = chr(ord(col) + 5)
            text_pos_right_row = row + 2 + 3
            text_pos_right_col = chr(ord(col) + 6)
        elif (2 == idx % 3):
            pic_pos_row = row
            pic_pos_col = chr(ord(col) + 8)
            text_pos_left_row = row + 1
            text_pos_left_col = chr(ord(col) + 9)
            text_pos_right_row = row + 2 + 3
            text_pos_right_col = chr(ord(col) + 10)
    elif (idx >= 3 and idx < 6):
        if (0 == idx % 3):
            pic_pos_row = row + 8
            pic_pos_col = col
            text_pos_left_row = row + 8 + 1
            text_pos_left_col = chr(ord(col) + 1)
            text_pos_right_row = row + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 2)
        elif (1 == idx % 3):
            pic_pos_row = row + 8
            pic_pos_col = chr(ord(col) + 4)
            text_pos_left_row = row + 8 + 1
            text_pos_left_col = chr(ord(col) + 5)
            text_pos_right_row = row + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 6)
        elif (2 == idx % 3):
            pic_pos_row = row + 8
            pic_pos_col = chr(ord(col) + 8)
            text_pos_left_row = row + 8 + 1
            text_pos_left_col = chr(ord(col) + 9)
            text_pos_right_row = row + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 10)
    elif (idx >= 6):
        if (0 == idx % 3):
            pic_pos_row = row + 8 + 8
            pic_pos_col = col
            text_pos_left_row = row + 8 + 8 + 1
            text_pos_left_col = chr(ord(col) + 1)
            text_pos_right_row = row + 8 + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 2)
        elif (1 == idx % 3):
            pic_pos_row = row + 8 + 8
            pic_pos_col = chr(ord(col) + 4)
            text_pos_left_row = row + 8 + 8 + 1
            text_pos_left_col = chr(ord(col) + 5)
            text_pos_right_row = row + 8 + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 6)
        elif (2 == idx % 3):
            pic_pos_row = row + 8 + 8
            pic_pos_col = chr(ord(col) + 8)
            text_pos_left_row = row + 8 + 8 + 1
            text_pos_left_col = chr(ord(col) + 9)
            text_pos_right_row = row + 8 + 8 + 2 + 3
            text_pos_right_col = chr(ord(col) + 10)

    text_start_pos = text_pos_left_col + str(text_pos_left_row)
    text_end_pos = text_pos_right_col + str(text_pos_right_row)
    ws.merge_cells(text_start_pos + ':' + text_end_pos)
    cs = ws[text_start_pos]
    cs.font = font
    cs.alignment = alignment
    cs.value = '             ' + str(c)

    pic_pos = pic_pos_col + str(pic_pos_row)
    pic = Image(pic_path)
    ws.add_image(pic, pic_pos)


def createTag(boxNumber,subBoxNumber, workSeq):
    box = gsBox.objects.get(boxNumber=boxNumber)
    if subBoxNumber:
        subBox = gsSubBox.objects.get(box=box, subBoxNumber=int(subBoxNumber))
        work = gsWork.objects.get(box=box, workSeq=workSeq, subBox=subBox)
    else:
        work = gsWork.objects.get(box=box, workSeq=workSeq)

    workName = work.workName
    # wts = gsWorkThing.objects.filter(work=work)
    # specialThingIDList = wts.values_list('thing', flat=True)
    # specialSerialNumberList = gsThing.objects.filter(id__in=specialThingIDList).values_list('serialNumber', flat=True)

    thing_set = gsThing.objects.filter(work__in=work)
    productTypeCode = box.productType
    productType = gsProperty.objects.get(project='实物类型', code=productTypeCode)

    if productType.type == u'金银锭类':
        rs = gsDing.objects.filter(thing__in=thing_set)
    elif productType.type == u'金银币章类':
        rs = gsBiZhang.objects.filter(thing__in=thing_set)
    elif productType.type == u'银元类':
        rs = gsYinYuan.objects.filter(thing__in=thing_set)
    elif productType.type == u'金银工艺品类':
        rs = gsGongYiPin.objects.filter(thing__in=thing_set)

    global tagDir
    tagDir = os.path.join(tagRootDir, str(boxNumber))
    if not os.path.exists(tagDir):
        os.mkdir(tagDir)

    global tempDir
    tempDir = os.path.join(tagDir, 'temp')
    if not os.path.exists(tempDir):
        os.mkdir(tempDir)

    xlFileName = u'{0}_标签.xlsx'.format(workName)
    xlFilePath = os.path.join(tagDir, xlFileName)
    if not os.path.exists(xlFilePath):
        # xlTmplateFileName = 'tag_template.xlsx'
        # xlTmplateFilePath = os.path.join(tagDir, xlTmplateFileName)
        # wb = load_workbook(xlTmplateFilePath)
        wb = Workbook()
        ws = wb.active
        # A workbook is always created with at least one worksheet.
        # You can get it by using the openpyxl.workbook.Workbook.active() property

        # r = rs[0]
        # c1 = r.serialNumber
        # c2 = c1+'-'+boxNumber
        # idx = 0
        # tag_pic = createQRCode(c2)
        # writeDataIntoXLS(ws, tag_pic, c1, idx)

        idx = 0
        start_col = 'C'
        start_row = 0
        for r in rs:
            # insert background picture
            if (0 == idx % 9):
                page = int(math.floor(idx / 9))
                row = start_row + page * 24
                # filePath = os.path.join(tagRootDir, 'background.png')
                # pic = Image(filePath)
                # ws.add_image(pic, start_col+str(row))

            c1 = r.serialNumber
            # c2 = c1+'-'+boxNumber
            c2 = c1
            tag_pic = createQRCode(c2)
            c1s = c1.split('-')
            c1 = c1s[0] + '-' + c1s[1] + '-' + c1s[2] + '-\n            ' + c1s[3]
            writeDataIntoXLS(ws, tag_pic, c1, row + 1, start_col, idx)

            idx = idx + 1

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_view.view = 'pageBreakPreview'
        # 单位: 英寸 inches. 1英寸==2.53CM
        ws.page_margins.left = 4.40 / 2.53
        ws.page_margins.right = 0.10 / 2.53
        ws.page_margins.top = 5.40 / 2.53
        ws.page_margins.bottom = 4.55 / 2.53
        ws.page_margins.header = 1.29 / 2.53
        ws.page_margins.footer = 1.29 / 2.53
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['J'].width = 5

        wb.save(xlFilePath)

        '''zipFileName = u'{0}_标签.zip'.format(workName)
        zipFilePath = os.path.join(tagDir, zipFileName)
        f = zipfile.ZipFile(zipFilePath, 'w' ,zipfile.ZIP_DEFLATED) 
        f.write(xlFilePath, xlFileName) 
        f.close()'''

        utils.deleteDir(tempDir)

def boxTag(boxNumber,userName,subBoxNumber):
    global tagDir
    tagDir = os.path.join(tagRootDir, str(boxNumber))
    if not os.path.exists(tagDir):
        os.mkdir(tagDir)

    global tempDir
    tempDir = os.path.join(tagDir, 'boxQR')
    if not os.path.exists(tempDir):
        os.mkdir(tempDir)

    if subBoxNumber == '':
        xlFileName = u'{0}号箱标签.xlsx'.format(boxNumber)
        xlFilePath = os.path.join(tagDir, xlFileName)
    else:
        xlFileName = u'{0}-{1}号箱标签.xlsx'.format(boxNumber, subBoxNumber)
        xlFilePath = os.path.join(tagDir, xlFileName)

    if not os.path.exists(xlFilePath):
        t = int(time.time() * 1000)
        ti = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        if subBoxNumber == '':
            data = ''.join([str(boxNumber),'号箱','|','实物分发岗位：',userName,'|',str(t)])
            gsBox.objects.filter(boxNumber=boxNumber).update(txtQR=data)
            txt = ''.join([str(boxNumber),'号箱','|','实物分发岗位：',userName,'|',str(ti)])
        else:
            data = ''.join([str(boxNumber), '-',str(subBoxNumber),'号箱','|','实物分发岗位：', userName,'|',str(t)])
            box = gsBox.objects.get(boxNumber=boxNumber)
            gsSubBox.objects.filter(box=box, subBoxNumber=subBoxNumber).update(txtQR=data)
            txt = ''.join([str(boxNumber), '-', str(subBoxNumber), '号箱', '|', '实物分发岗位：', userName, '|', str(ti)])

        img = qrcode.make(data=data)
        picName = xlFileName.replace('.xlsx','.png')
        filePath = os.path.join(tempDir, picName)
        img.resize((256, 256)).save(filePath)

        # 将图片插入高excel中
        wb = Workbook()
        ws = wb.active

        start_col = 'C'
        start_row = 2

        # QRCode picture and Text
        font = Font(name=u'Times New Roman', size=13, bold=True)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        pic_pos_row = start_row
        pic_pos_col = start_col
        text_pos_left_row = start_row + 13
        text_pos_left_col = chr(ord(start_col))
        text_pos_right_row = start_row + 16
        text_pos_right_col = chr(ord(start_col)+3)

        text_start_pos = text_pos_left_col + str(text_pos_left_row)
        text_end_pos = text_pos_right_col + str(text_pos_right_row)
        ws.merge_cells(text_start_pos + ':' + text_end_pos)
        cs = ws[text_start_pos]
        cs.font = font
        cs.alignment = alignment
        cs.value = txt

        pic_pos = pic_pos_col + str(pic_pos_row)
        pic = Image(filePath)
        ws.add_image(pic, pic_pos)

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_view.view = 'pageBreakPreview'
        # 单位: 英寸 inches. 1英寸==2.53CM
        ws.page_margins.left = 4.40 / 2.53
        ws.page_margins.right = 4.40 / 2.53
        ws.page_margins.top = 5.40 / 2.53
        #ws.page_margins.bottom = 6.40 / 2.53
        ws.page_margins.header = 4.40 / 2.53
        ws.page_margins.footer = 4.40 / 2.53
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['J'].width = 5

        wb.save(xlFilePath)
        #utils.deleteDir(tempDir)
    return True


def packageTag(boxNumber,subBoxNumber,packageList):
    global tagDir
    tagDir = os.path.join(tagRootDir, str(boxNumber))
    if not os.path.exists(tagDir):
        os.mkdir(tagDir)

    global tempDir
    tempDir = os.path.join(tagDir, 'boxQR')
    if not os.path.exists(tempDir):
        os.mkdir(tempDir)

    if subBoxNumber == '':
        xlFileName = u'{0}号箱包标签.xlsx'.format(boxNumber)
        xlFilePath = os.path.join(tagDir, xlFileName)
    else:
        xlFileName = u'{0}-{1}号箱包标签.xlsx'.format(boxNumber, subBoxNumber)
        xlFilePath = os.path.join(tagDir, xlFileName)
    if not os.path.exists(xlFilePath):
        wb = Workbook()
        ws = wb.active

        idx = 0
        start_col = 'C'
        start_row = 0
        for packageNo in packageList:
            # insert background picture
            if (0 == idx % 9):
                page = int(math.floor(idx / 9))
                row = start_row + page * 24

            tag_pic = createQRCode(packageNo)

            writeDataIntoXLS(ws, tag_pic, packageNo, row + 1, start_col, idx)

            idx = idx + 1

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_view.view = 'pageBreakPreview'
        # 单位: 英寸 inches. 1英寸==2.53CM
        ws.page_margins.left = 4.40 / 2.53
        ws.page_margins.right = 0.10 / 2.53
        ws.page_margins.top = 5.40 / 2.53
        ws.page_margins.bottom = 4.55 / 2.53
        ws.page_margins.header = 1.29 / 2.53
        ws.page_margins.footer = 1.29 / 2.53
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['J'].width = 5

        wb.save(xlFilePath)

        utils.deleteDir(tempDir)